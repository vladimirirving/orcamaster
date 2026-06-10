import io
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from app.database import get_db
from app.dependencies import get_current_user
from app.models.grupo import Grupo
from app.models.item import Item
from app.models.obra import Obra
from app.models.usuario import Usuario
from app.models.versao import Versao
from app.schemas.curva_abc import CurvaAbcData, CurvaAbcItem

router = APIRouter(tags=["curva_abc"])

FAIXA_FILLS = {
    "A": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
    "B": PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid"),
    "C": PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid"),
}


async def _get_versao(versao_id: int, current_user: Usuario, db: AsyncSession) -> Versao:
    result = await db.execute(
        select(Versao)
        .join(Obra, Versao.obra_id == Obra.id)
        .where(Versao.id == versao_id, Obra.empresa_id == current_user.empresa_id)
    )
    versao = result.scalar_one_or_none()
    if versao is None:
        raise HTTPException(status_code=404, detail="Versão não encontrada")
    return versao


async def _get_itens(versao_id: int, db: AsyncSession) -> list[Item]:
    result = await db.execute(
        select(Item)
        .join(Grupo, Item.grupo_id == Grupo.id)
        .where(Grupo.versao_id == versao_id)
        .options(selectinload(Item.grupo), selectinload(Item.composicao))
        .order_by(Item.total.desc())
    )
    return result.scalars().all()


def _calcular_abc(versao: Versao, itens: list) -> CurvaAbcData:
    from decimal import Decimal, ROUND_HALF_UP

    total_versao = Decimal(str(versao.total_sem_bdi or 0))

    if total_versao == 0:
        return CurvaAbcData(total_versao=str(versao.total_sem_bdi or "0.00"), itens=[])

    itens_com_valor = [i for i in itens if float(i.total) > 0]

    resultado = []
    acumulado = Decimal("0")
    for rank, item in enumerate(itens_com_valor, start=1):
        total_item = Decimal(str(item.total))
        participacao = (total_item / total_versao * 100).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        acumulado = (acumulado + participacao).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        faixa = "A" if acumulado <= 80 else "B" if acumulado <= 95 else "C"
        resultado.append(CurvaAbcItem(
            rank=rank,
            grupo_nome=item.grupo.nome,
            descricao=item.composicao.descricao if item.composicao else "",
            unidade=item.unidade,
            quantidade=str(item.quantidade),
            total=str(round(item.total, 2)),
            participacao_pct=float(participacao),
            acumulado_pct=float(acumulado),
            faixa=faixa,
        ))

    return CurvaAbcData(
        total_versao=str(versao.total_sem_bdi),
        itens=resultado,
    )


def _build_xlsx(data: CurvaAbcData) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Curva ABC"

    headers = ["#", "Grupo", "Descrição", "Unidade", "Quantidade", "Total (R$)", "Part%", "Acum%", "Faixa"]
    header_font = Font(bold=True)
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font

    col_widths = [5, 20, 40, 10, 15, 18, 10, 10, 8]
    for i, w in enumerate(col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    for item in data.itens:
        ws.append([
            item.rank,
            item.grupo_nome,
            item.descricao,
            item.unidade,
            float(item.quantidade),
            float(item.total),
            item.participacao_pct,
            item.acumulado_pct,
            item.faixa,
        ])
        fill = FAIXA_FILLS[item.faixa]
        for col in range(1, 10):
            ws.cell(row=ws.max_row, column=col).fill = fill

    # Linha de total (only if there are items)
    if data.itens:
        total_row = ws.max_row + 1
        ws.cell(row=total_row, column=1, value="Total").font = Font(bold=True)
        ws.cell(row=total_row, column=6, value=float(data.total_versao)).font = Font(bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


@router.get("/versoes/{versao_id}/curva-abc", response_model=CurvaAbcData)
async def get_curva_abc(
    versao_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    versao = await _get_versao(versao_id, current_user, db)
    itens = await _get_itens(versao_id, db)
    return _calcular_abc(versao, itens)


@router.get("/versoes/{versao_id}/curva-abc/export")
async def export_curva_abc(
    versao_id: int,
    current_user: Usuario = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    versao = await _get_versao(versao_id, current_user, db)
    itens = await _get_itens(versao_id, db)
    data = _calcular_abc(versao, itens)
    xlsx_bytes = _build_xlsx(data)

    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="curva-abc-v{versao_id}.xlsx"'
        },
    )
