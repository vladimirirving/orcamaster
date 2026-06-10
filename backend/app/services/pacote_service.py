import logging
import os
import zipfile
from datetime import datetime
from io import BytesIO
from pathlib import Path
from typing import Optional

logger = logging.getLogger(__name__)

import openpyxl
from openpyxl.styles import Font, PatternFill
from fastapi import HTTPException
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload

from app.config import settings
from app.database import AsyncSessionLocal
from app.models.bdi import BDI
from app.models.cronograma_linha import CronogramaLinha
from app.models.grupo import Grupo
from app.models.item import Item
from app.models.pacote_job import PacoteJob
from app.models.versao import Versao
from app.services.proposta_pdf import gerar_pdf_bytes


async def gerar_planilha_bytes(versao_id: int, db: AsyncSession) -> bytes:
    versao_result = await db.execute(select(Versao).where(Versao.id == versao_id))
    versao = versao_result.scalar_one()

    bdi_result = await db.execute(select(BDI).where(BDI.versao_id == versao_id))
    bdi: Optional[BDI] = bdi_result.scalar_one_or_none()

    grupos_result = await db.execute(
        select(Grupo)
        .where(Grupo.versao_id == versao_id)
        .options(selectinload(Grupo.itens).selectinload(Item.composicao))
        .order_by(Grupo.ordem)
    )
    grupos = grupos_result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha Orçamentária"

    bold = Font(bold=True)
    group_fill = PatternFill("solid", fgColor="D0D0D0")
    total_fill = PatternFill("solid", fgColor="B0B0B0")

    ws.append(["Código", "Descrição", "Un", "Qtd", "Preço Unit. (R$)", "Total (R$)"])
    for cell in ws[1]:
        cell.font = bold

    for grupo in grupos:
        indent = "  " if grupo.pai_id is not None else ""
        ws.append([grupo.codigo or "", indent + grupo.nome, "", "", "", ""])
        for cell in ws[ws.max_row]:
            cell.fill = group_fill
            cell.font = bold

        group_total = 0.0
        for item in sorted(grupo.itens, key=lambda x: x.ordem):
            descricao = item.composicao.descricao if item.composicao else "—"
            codigo = item.composicao.codigo if item.composicao else ""
            item_total = float(item.total)
            group_total += item_total
            ws.append([
                codigo,
                descricao,
                item.unidade,
                float(item.quantidade),
                float(item.preco_unitario_sem_bdi) if item.preco_unitario_sem_bdi else 0.0,
                item_total,
            ])

        ws.append(["", f"  Subtotal — {grupo.nome}", "", "", "", group_total])
        for cell in ws[ws.max_row]:
            cell.font = bold

    total_sem = float(versao.total_sem_bdi)
    total_com = float(versao.total_com_bdi)

    ws.append(["", "SUBTOTAL SEM BDI", "", "", "", total_sem])
    for cell in ws[ws.max_row]:
        cell.font = bold

    if bdi:
        bdi_pct = float(bdi.bdi_composto) * 100
        ws.append(["", f"BDI ({bdi_pct:.2f}%)", "", "", "", total_com - total_sem])

    ws.append(["", "TOTAL GERAL", "", "", "", total_com])
    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = bold

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 18

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def gerar_cronograma_bytes(versao_id: int, db: AsyncSession) -> Optional[bytes]:
    versao_result = await db.execute(select(Versao).where(Versao.id == versao_id))
    versao = versao_result.scalar_one()

    if not versao.cronograma_inicio:
        return None

    inicio = versao.cronograma_inicio   # "YYYY-MM"
    fim = versao.cronograma_fim or inicio

    meses: list[str] = []
    y, m = int(inicio[:4]), int(inicio[5:7])
    ey, em = int(fim[:4]), int(fim[5:7])
    while (y, m) <= (ey, em):
        meses.append(f"{y:04d}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1

    linhas_result = await db.execute(
        select(CronogramaLinha)
        .join(Item, CronogramaLinha.item_id == Item.id)
        .join(Grupo, Item.grupo_id == Grupo.id)
        .where(Grupo.versao_id == versao_id)
        .options(selectinload(CronogramaLinha.item).selectinload(Item.composicao))
        .order_by(Grupo.ordem, Item.ordem)
    )
    linhas = linhas_result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cronograma"

    bold = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="B0B0B0")

    ws.append(["Item", "Descrição", "Un", "Qtd", "Total (R$)"] + meses)
    for cell in ws[1]:
        cell.font = bold

    monthly_totals = {mes: 0.0 for mes in meses}

    for linha in linhas:
        item = linha.item
        descricao = item.composicao.descricao if item.composicao else "—"
        total_val = float(item.total)
        row = [
            item.composicao.codigo if item.composicao else "",
            descricao,
            item.unidade,
            float(item.quantidade),
            total_val,
        ]
        for mes in meses:
            pct = linha.distribuicao_json.get(mes, 0.0)
            val = round(total_val * pct / 100.0, 2)
            monthly_totals[mes] += val
            row.append(val)
        ws.append(row)

    item_total_sum = round(sum(float(linha.item.total) for linha in linhas), 2)
    totals_row = ["", "TOTAL MENSAL", "", "", item_total_sum]
    totals_row += [round(monthly_totals[mes], 2) for mes in meses]
    ws.append(totals_row)
    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = bold

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 6
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 14

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def processar_pacote(job_id: int, versao_id: int, _db_factory=None) -> None:
    """Gera ZIP com proposta PDF + planilha XLSX + cronograma XLSX. Chamado via BackgroundTasks."""
    factory = _db_factory or AsyncSessionLocal

    async with factory() as db:
        result = await db.execute(select(PacoteJob).where(PacoteJob.id == job_id))
        job = result.scalar_one()
        job.status = "processando"
        await db.commit()

    try:
        async with factory() as db:
            pdf_bytes: Optional[bytes] = None
            try:
                pdf_bytes = await gerar_pdf_bytes(versao_id, db)
            except HTTPException:
                pass

            planilha_bytes = await gerar_planilha_bytes(versao_id, db)
            cronograma_bytes = await gerar_cronograma_bytes(versao_id, db)

        zip_buf = BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            if pdf_bytes:
                zf.writestr("proposta.pdf", pdf_bytes)
            zf.writestr("planilha.xlsx", planilha_bytes)
            if cronograma_bytes:
                zf.writestr("cronograma.xlsx", cronograma_bytes)

        pacotes_dir = Path(settings.pacotes_dir)
        pacotes_dir.mkdir(mode=0o700, parents=True, exist_ok=True)
        filename = f"pacote-v{versao_id}-{job_id}.zip"
        file_path = pacotes_dir / filename
        fd = os.open(str(file_path), os.O_WRONLY | os.O_CREAT | os.O_TRUNC, 0o600)
        try:
            with os.fdopen(fd, "wb") as f:
                f.write(zip_buf.getvalue())
        except Exception:
            os.close(fd)
            raise

        async with factory() as db:
            result = await db.execute(select(PacoteJob).where(PacoteJob.id == job_id))
            job = result.scalar_one()
            job.status = "pronto"
            job.url_download = filename
            job.gerado_em = datetime.utcnow()
            await db.commit()

    except Exception:
        logger.exception("processar_pacote job_id=%s versao_id=%s failed", job_id, versao_id)
        async with factory() as db:
            result = await db.execute(select(PacoteJob).where(PacoteJob.id == job_id))
            job = result.scalar_one()
            job.status = "erro"
            job.erro_mensagem = "Erro interno ao gerar o pacote. Contate o suporte."
            await db.commit()
