import io
import zipfile
from collections import defaultdict
from decimal import Decimal
from typing import Optional

import openpyxl
from fastapi import HTTPException
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.composicao import Composicao
from app.models.grupo import Grupo
from app.models.item import Item
from app.models.versao import Versao
from app.schemas.planilha_import import ImportarPlanilhaResult
from app.services.totais_service import recalc_totais_versao


def gerar_template_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Planilha"

    bold = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    example_fill = PatternFill("solid", fgColor="F5F5F5")
    example_font = Font(italic=True, color="888888")

    headers = ["grupo", "subgrupo", "codigo_composicao", "descricao", "unidade", "quantidade"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = bold
        cell.fill = header_fill

    ws.append(["Terraplenagem", "", "94966", "Escavação mecânica de vala", "M3", 12000])
    for cell in ws[2]:
        cell.font = example_font
        cell.fill = example_fill

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 36
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


async def importar_planilha(
    versao_id: int,
    empresa_id: int,
    conteudo: bytes,
    db: AsyncSession,
) -> ImportarPlanilhaResult:
    try:
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
    except (zipfile.BadZipFile, KeyError, Exception) as exc:
        raise HTTPException(status_code=400, detail="Arquivo inválido. Envie um arquivo .xlsx válido.") from exc
    try:
        ws = wb.active
        all_rows = list(ws.iter_rows(values_only=True)) if ws is not None else []
    finally:
        wb.close()

    if not all_rows:
        return ImportarPlanilhaResult(grupos_criados=0, itens_criados=0, itens_sem_composicao=0)

    header = [str(h).strip().lower() if h is not None else "" for h in all_rows[0]]

    def cell(row, col_name: str) -> str:
        try:
            idx = header.index(col_name)
            val = row[idx]
            return str(val).strip() if val is not None else ""
        except ValueError:
            return ""

    r_max = await db.execute(
        select(func.coalesce(func.max(Grupo.ordem), -1)).where(Grupo.versao_id == versao_id)
    )
    next_ordem = r_max.scalar() + 1

    grupo_cache: dict[str, Grupo] = {}
    subgrupo_cache: dict[tuple[str, str], Grupo] = {}
    item_ordem: dict[int, int] = defaultdict(int)

    grupos_criados = 0
    itens_criados = 0
    itens_sem_composicao = 0

    for raw_row in all_rows[1:]:
        nome_grupo = cell(raw_row, "grupo")
        if not nome_grupo:
            continue

        if nome_grupo not in grupo_cache:
            g = Grupo(versao_id=versao_id, nome=nome_grupo, ordem=next_ordem)
            db.add(g)
            await db.flush()
            grupo_cache[nome_grupo] = g
            next_ordem += 1
            grupos_criados += 1

        parent = grupo_cache[nome_grupo]

        nome_sub = cell(raw_row, "subgrupo")
        if nome_sub:
            key = (nome_grupo, nome_sub)
            if key not in subgrupo_cache:
                sg = Grupo(versao_id=versao_id, nome=nome_sub,
                           pai_id=parent.id, ordem=next_ordem)
                db.add(sg)
                await db.flush()
                subgrupo_cache[key] = sg
                next_ordem += 1
                grupos_criados += 1
            item_parent = subgrupo_cache[key]
        else:
            item_parent = parent

        codigo = cell(raw_row, "codigo_composicao")
        comp: Optional[Composicao] = None
        if codigo:
            r = await db.execute(
                select(Composicao).where(
                    Composicao.codigo == codigo,
                    or_(Composicao.empresa_id.is_(None),
                        Composicao.empresa_id == empresa_id),
                ).limit(1)
            )
            comp = r.scalar_one_or_none()

        qtd_raw = cell(raw_row, "quantidade").replace(",", ".")
        try:
            quantidade = Decimal(qtd_raw) if qtd_raw else Decimal("0")
        except Exception:
            quantidade = Decimal("0")

        unidade = cell(raw_row, "unidade") or "UN"
        requer_revisao = comp is None and bool(codigo)

        db.add(Item(
            grupo_id=item_parent.id,
            composicao_id=comp.id if comp else None,
            ordem=item_ordem[item_parent.id],
            quantidade=quantidade,
            unidade=unidade,
            preco_unitario_sem_bdi=comp.preco_unitario if comp else None,
            requer_revisao=requer_revisao,
        ))

        itens_criados += 1
        if requer_revisao:
            itens_sem_composicao += 1
        item_ordem[item_parent.id] += 1

    await db.flush()
    await recalc_totais_versao(versao_id, db)
    await db.commit()

    return ImportarPlanilhaResult(
        grupos_criados=grupos_criados,
        itens_criados=itens_criados,
        itens_sem_composicao=itens_sem_composicao,
    )
